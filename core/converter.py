"""
Excel Converter Module
Handles conversion of extracted PDF data to Excel format
"""

import pandas as pd
import os
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)

class ExcelConverter:
    """
    Converts extracted PDF data to Excel format with formatting
    """
    
    def __init__(self):
        """Initialize Excel Converter"""
        self.workbook = None
        self.writer = None
    
    def _create_header_style(self):
        """Create header styling for Excel"""
        return {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _create_data_style(self):
        """Create data cell styling for Excel"""
        return {
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _apply_styling(self, worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
        """Apply styling to a range of cells"""
        header_style = self._create_header_style()
        data_style = self._create_data_style()
        
        # Apply header styling to first row
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
        
        # Apply data styling to remaining rows
        for row in range(start_row + 1, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_style['alignment']
                cell.border = data_style['border']
    
    def _auto_adjust_column_widths(self, worksheet):
        """Automatically adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def convert_to_excel(self, extraction_results: Dict[str, Any], output_path: str) -> bool:
        """
        Convert extracted PDF data to Excel format - Single consolidated file
        Uses hybrid approach: text-based extraction with OCR fallback
        
        Args:
            extraction_results (Dict[str, Any]): Results from PDF extraction
            output_path (str): Path for output Excel file
            
        Returns:
            bool: True if conversion successful, False otherwise
        """
        try:
            tables = extraction_results.get('tables', [])
            if not tables:
                logger.warning("No tables found to convert")
                return False
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            logger.info(f"📊 Converting to Excel: {output_path}")
            
            # Combine all tables into one consolidated DataFrame
            consolidated_data = []
            
            for table in tables:
                # Convert DataFrame to list of dictionaries for easier processing
                table_data = table.to_dict('records')
                consolidated_data.extend(table_data)
            
            # Create the consolidated DataFrame
            if consolidated_data:
                consolidated_df = pd.DataFrame(consolidated_data)
                
                # Reorder columns to put Page and Table_Index first if they exist
                column_order = []
                if 'Page' in consolidated_df.columns:
                    column_order.append('Page')
                if 'Table_Index' in consolidated_df.columns:
                    column_order.append('Table_Index')
                
                # Add remaining columns
                for col in consolidated_df.columns:
                    if col not in column_order:
                        column_order.append(col)
                
                consolidated_df = consolidated_df[column_order]
                
                # Use openpyxl for better control and formatting
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                
                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "PDF_Data"
                
                # Write headers
                headers = list(consolidated_df.columns)
                ws.append(headers)
                
                # Write data rows
                for _, row in consolidated_df.iterrows():
                    ws.append(list(row))
                
                # Apply styling to headers
                header_style = self._create_header_style()
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_style['font']
                    cell.fill = header_style['fill']
                    cell.alignment = header_style['alignment']
                    cell.border = header_style['border']
                
                # Apply styling to data rows
                data_style = self._create_data_style()
                for row in range(2, len(consolidated_df) + 2):
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = data_style['alignment']
                        cell.border = data_style['border']
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    max_length = 0
                    for row in range(1, len(consolidated_df) + 2):
                        cell = ws.cell(row=row, column=col)
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width with some padding
                    adjusted_width = min(max_length + 2, 50)  # Max width of 50
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # Save the workbook
                wb.save(output_path)
                
                logger.info(f"✅ Created consolidated Excel file with {len(consolidated_df)} rows")
            else:
                logger.warning("No data to consolidate")
                return False
            
            logger.info(f"🎉 Excel conversion complete: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error converting to Excel: {e}")
            return False
    
    def create_summary_sheet(self, extraction_results: Dict[str, Any], output_path: str) -> bool:
        """
        Create a summary sheet with extraction metadata
        
        Args:
            extraction_results (Dict[str, Any]): Results from PDF extraction
            output_path (str): Path for output Excel file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create summary data
            summary_data = {
                'Property': [
                    'PDF File',
                    'Extraction Method',
                    'PDF Type',
                    'Total Pages',
                    'Total Sections',
                    'Total Rows',
                    'Processing Time'
                ],
                'Value': [
                    os.path.basename(extraction_results.get('pdf_path', 'Unknown')),
                    extraction_results.get('extraction_method', 'Unknown'),
                    'Text-based' if extraction_results.get('is_text_based', False) else 'Image-based',
                    extraction_results.get('total_pages', 0),
                    extraction_results.get('total_tables', 0),
                    extraction_results.get('total_rows', 0),
                    'N/A'  # Could be calculated if timing is tracked
                ]
            }
            
            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Append to existing Excel file
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Get the worksheet for styling
                worksheet = writer.sheets['Summary']
                
                # Apply styling to summary sheet
                self._apply_styling(
                    worksheet, 
                    start_row=1, 
                    end_row=len(summary_df) + 1, 
                    start_col=1, 
                    end_col=len(summary_df.columns)
                )
                
                # Auto-adjust column widths
                self._auto_adjust_column_widths(worksheet)
            
            logger.info("✅ Summary sheet created")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error creating summary sheet: {e}")
            return False 